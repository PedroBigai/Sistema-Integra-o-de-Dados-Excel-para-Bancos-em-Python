import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook


# Inicialize o aplicativo Firebase com as credenciais
cred = credentials.Certificate(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa_preencher_sit/flutter-firebase-db-c5531-firebase-adminsdk-p516p-30bde2e024.json"
)
firebase_admin.initialize_app(cred)


planilha = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Rolo.xlsx"
)  # Substitua pelo caminho correto
planilha_rolo = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Tabela Jô Decorações - ROLO + DV + SCREEN + BK (1).xlsx"
)  # Substitua pelo caminho correto
ws_rolo = planilha["Table 1"]
ws_rolo2 = planilha_rolo["Table 2"]

for row in ws_rolo.iter_rows(
    min_row=3, values_only=True, min_col=0, max_col=10, max_row=39
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[6]
    preco_prazo = row[8]
    modelo = "Rolô"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)
    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


for row in ws_rolo2.iter_rows(
    min_row=9, values_only=True, min_col=0, max_col=10, max_row=18
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "Rolô"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_rolo2.iter_rows(
    min_row=20, values_only=True, min_col=0, max_col=10, max_row=32
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "Rolô"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_rolo2.iter_rows(
    min_row=2, values_only=True, min_col=0, max_col=10, max_row=7
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "Dv Screen"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


planilha_romana = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Romana.xlsx"
)  # Substitua pelo caminho correto
ws_romana = planilha_romana["Table 1"]

for row in ws_romana.iter_rows(
    min_row=4, values_only=True, min_col=0, max_col=10, max_row=30
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[4]
    preco_vista = row[6]
    preco_prazo = row[8]
    modelo = "Romana"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_romana.iter_rows(
    min_row=32, values_only=True, min_col=0, max_col=10, max_row=41
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[4]
    preco_vista = row[6]
    preco_prazo = row[8]
    modelo = "Romana"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_romana.iter_rows(
    min_row=43, values_only=True, min_col=0, max_col=10, max_row=54
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[4]
    preco_vista = row[6]
    preco_prazo = row[8]
    modelo = "Romana"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

planilha_ph = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/PH.xlsx"
)  # Substitua pelo caminho correto
ws_ph = planilha_ph["Table 1"]
for row in ws_ph.iter_rows(
    min_row=4, values_only=True, min_col=0, max_col=7, max_row=10
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    modelo = "PH"
    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_ph.iter_rows(
    min_row=12, values_only=True, min_col=0, max_col=7, max_row=14
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    modelo = "PH"
    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}, código:{codigo}, largura: {largura}, R${preco_vista}, {preco_prazo}"
    )

for row in ws_ph.iter_rows(
    min_row=16, values_only=True, min_col=0, max_col=7, max_row=19
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    modelo = "PH Standard"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_ph.iter_rows(
    min_row=21, values_only=True, min_col=0, max_col=7, max_row=24
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "PH Monocomando"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


planilha_dv = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Rolo DV Screen.xlsx"
)  # Substitua pelo caminho correto
ws_dv = planilha_dv["Table 2"]

for row in ws_dv.iter_rows(
    min_row=2, values_only=True, min_col=0, max_col=10, max_row=7
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "DV Screen"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)
    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_dv.iter_rows(
    min_row=9, values_only=True, min_col=0, max_col=10, max_row=18
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "DV Screen"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_dv.iter_rows(
    min_row=20, values_only=True, min_col=0, max_col=10, max_row=32
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    largura = row[2]
    preco_vista = row[4]
    preco_prazo = row[6]
    modelo = "DV Screen"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


planilha3 = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Pv.xlsx"
)  # Substitua pelo caminho correto
ws_pv = planilha3["Table 1"]

for row in ws_pv.iter_rows(
    min_row=4, values_only=True, min_col=0, max_col=7, max_row=36
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "PV"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    preco_vista is int
    preco_prazo is int

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


for row in ws_pv.iter_rows(
    min_row=38, values_only=True, min_col=0, max_col=7, max_row=40
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "PV"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    preco_vista is int
    preco_prazo is int

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


planilha4 = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/PV BK.xlsx"
)  # Substitua pelo caminho correto
ws_pv_bk = planilha4["Table 1"]

for row in ws_pv_bk.iter_rows(
    min_row=4, values_only=True, min_col=0, max_col=7, max_row=19
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    codigo = row[1]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "PV BK"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())
    preco_vista is int
    preco_prazo is int

    dados_produto = {
        "Modelo": f"{modelo}",
        "Tecido": f"{tecido}",
        "codigo": f"{codigo}",
        "largura": largura,
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("produtos").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, tecido {tecido}\n, código:{codigo}\n, largura: {largura}\n, R${preco_vista}\n, {preco_prazo}\n"
    )


planilha_acess = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Acess Rolo.xlsx"
)  # Substitua pelo caminho correto
ws_rolo_acess = planilha_acess["Table 2"]

for row in ws_rolo_acess.iter_rows(
    min_row=2, values_only=True, min_col=0, max_col=6, max_row=21
):  # Substitua o número de linhas máximo conforme necessário
    acess = row[0]
    preco_vista = row[2]
    preco_prazo = row[4]
    modelo = "ACESSORIOS ROLO"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_rolo_acess.iter_rows(
    min_row=23, values_only=True, min_col=0, max_col=6, max_row=28
):  # Substitua o número de linhas máximo conforme necessário
    acess = row[0]
    preco_vista = row[2]
    preco_prazo = row[4]
    modelo = "ACESSORIOS ROLO (MOTORIZAÇÃO)"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

planilha_romana_acess = load_workbook(
    "C:/Users/Pedro/OneDrive/Área de Trabalho/programas python/programa kivy/Acess Romana.xlsx"
)  # Substitua pelo caminho correto
ws_acess_romana = planilha_romana_acess["Table 2"]

for row in ws_acess_romana.iter_rows(
    min_row=2, values_only=True, min_col=0, max_col=6, max_row=20
):  # Substitua o número de linhas máximo conforme necessário
    acess = row[0]
    preco_vista = row[2]
    preco_prazo = row[4]
    modelo = "ACESSÓRIOS ROMANA"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_acess_romana.iter_rows(
    min_row=22, values_only=True, min_col=0, max_col=6, max_row=27
):  # Substitua o número de linhas máximo conforme necessário
    acess = row[0]
    preco_vista = row[2]
    preco_prazo = row[4]
    modelo = "ACESSÓRIOS ROMANA (MOTORIZAÇÃO)"

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_ph.iter_rows(
    min_row=26, values_only=True, min_col=0, max_col=7, max_row=28
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "ACESSÓRIOS PH"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )

for row in ws_pv_bk.iter_rows(
    min_row=21, values_only=True, min_col=0, max_col=7, max_row=41
):  # Substitua o número de linhas máximo conforme necessário
    tecido = row[0]
    preco_vista = row[5]
    preco_prazo = row[6]
    modelo = "PV BK (ACESSÓRIOS)"

    preco_vista = float(str(preco_vista).replace("R$", "").replace(",", ".").strip())
    preco_prazo = float(str(preco_prazo).replace("R$", "").replace(",", ".").strip())

    dados_produto = {
        "Modelo": f"{modelo}",
        "Acessório": f"{acess}",
        "preco_vista": preco_vista,
        "preco_prazo": preco_prazo,
    }

    # Adiciona os dados ao documento no Firestore
    # Obtenha uma referência para o banco de dados Firestore
    db = firestore.client()
    produto_ref = db.collection("acessorios").document()
    produto_ref.set(dados_produto)

    print(
        f"modelo: {modelo}\n, acessório {acess}\n, R${preco_vista}\n, {preco_prazo}\n"
    )
